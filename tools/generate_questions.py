import json
import posixpath
import random
import re
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DOC_DIR = ROOT / 'document'
OUT_JS_FILE = ROOT / 'src' / 'data' / 'questions.js'
OUT_JSON_FILE = ROOT / 'src' / 'data' / 'questions.json'
ASSET_DIR = ROOT / 'src' / 'assets' / 'questions'

# 出力先追加（D:\Project\denki）
DENKI_ROOT = ROOT.parent / 'denki'
DENKI_OUT_JS = DENKI_ROOT / 'src' / 'data' / 'questions.js'
DENKI_ASSET_DIR = DENKI_ROOT / 'assets' / 'questions'

QUESTION_LABEL_RE = re.compile(r'^\((\d+)\)$')
NUMBER_RE = re.compile(r'[-+]?\d[\d,]*(?:\.\d+)?(?:[eE][-+]?\d+)?')
SKIP_SHEETS = {'目次', '加工方法'}
RANDOM = random.Random(42)

NS = {
    'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'rel': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'pkg': 'http://schemas.openxmlformats.org/package/2006/relationships',
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
}


def norm(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float):
        return f"{value:.12g}"
    s = str(value).strip()
    return re.sub(r'\s+', ' ', s)


def safe_slug(text: str) -> str:
    t = re.sub(r'\s+', '_', text.strip())
    t = re.sub(r'[^0-9A-Za-zぁ-んァ-ン一-龯_\-]', '', t)
    return t[:60] or 'sheet'


def parse_sheet(ws):
    questions: Dict[int, Dict[str, object]] = {}

    for row in range(1, ws.max_row + 1):
        label = norm(ws.cell(row=row, column=1).value)
        q_text = norm(ws.cell(row=row, column=2).value)
        m = QUESTION_LABEL_RE.match(label)
        if not m or not q_text or len(q_text) < 5:
            continue
        idx = int(m.group(1))
        questions[idx] = {'text': q_text, 'row': row}

    answer_start = None
    for row in range(1, ws.max_row + 1):
        for col in range(1, min(ws.max_column, 40) + 1):
            if norm(ws.cell(row=row, column=col).value) == '答え':
                answer_start = row
                break
        if answer_start is not None:
            break

    answers: Dict[int, str] = {}
    if answer_start is not None:
        for row in range(answer_start + 1, ws.max_row + 1):
            col = 1
            while col <= ws.max_column:
                value = norm(ws.cell(row=row, column=col).value)
                m = QUESTION_LABEL_RE.match(value)
                if not m:
                    col += 1
                    continue

                qn = int(m.group(1))
                tokens: List[str] = []
                scan = col + 1
                while scan <= ws.max_column:
                    nxt = norm(ws.cell(row=row, column=scan).value)
                    if QUESTION_LABEL_RE.match(nxt):
                        break
                    if nxt:
                        tokens.append(nxt)
                    scan += 1

                if tokens and qn not in answers:
                    answers[qn] = ' '.join(tokens)

                col = scan

    return questions, answers


def parse_numeric_answer(answer_text: str) -> Optional[Tuple[float, str, float, float]]:
    m = NUMBER_RE.search(answer_text)
    if not m:
        return None

    token = m.group(0)
    try:
        value = float(token.replace(',', ''))
    except ValueError:
        return None

    unit = answer_text[m.end():].strip()
    abs_tol = max(abs(value) * 0.02, 1e-6)
    rel_tol = 0.02
    return value, unit, abs_tol, rel_tol


def create_numeric_choices(answer_text):
    parsed = parse_numeric_answer(answer_text)
    if not parsed:
        return None

    value, unit, _, _ = parsed
    factors = [0.5, 0.8, 1.2, 1.5, 2.0, 0.75, 1.25]
    RANDOM.shuffle(factors)

    distractors = []
    for factor in factors:
        d = value * factor
        if abs(value) > 1e-12 and abs(d - value) / abs(value) < 1e-9:
            continue
        distractors.append(d)
        if len(distractors) == 3:
            break

    if len(distractors) < 3:
        return None

    def fmt(v):
        av = abs(v)
        if av >= 1e6 or (0 < av < 1e-3):
            s = f"{v:.3e}"
        elif av >= 100:
            s = f"{v:.3f}".rstrip('0').rstrip('.')
        else:
            s = f"{v:.6g}"
        return f"{s}{(' ' + unit) if unit else ''}".strip()

    correct = answer_text
    raw_choices = [correct] + [fmt(x) for x in distractors]

    unique = []
    for c in raw_choices:
        if c not in unique:
            unique.append(c)
    if len(unique) < 4:
        return None

    choices = unique[:4]
    RANDOM.shuffle(choices)
    answer_idx = choices.index(correct)
    return choices, answer_idx


def create_fallback_choices(answer_text):
    choices = [
        answer_text,
        f"{answer_text} + 10%",
        f"{answer_text} - 10%",
        '別の値',
    ]
    RANDOM.shuffle(choices)
    return choices, choices.index(answer_text)


def _join(base: str, target: str) -> str:
    return posixpath.normpath(posixpath.join(posixpath.dirname(base), target))


def extract_images_from_xlsx(xlsx_path: Path, chapter_slug: str) -> Dict[str, List[dict]]:
    ASSET_DIR.mkdir(parents=True, exist_ok=True)
    result: Dict[str, List[dict]] = {}

    with zipfile.ZipFile(xlsx_path, 'r') as zf:
        def read_xml(path: str):
            try:
                return ET.fromstring(zf.read(path))
            except KeyError:
                return None

        workbook = read_xml('xl/workbook.xml')
        wb_rels = read_xml('xl/_rels/workbook.xml.rels')
        if workbook is None or wb_rels is None:
            return result

        rel_map = {
            rel.attrib.get('Id'): rel.attrib.get('Target')
            for rel in wb_rels.findall('pkg:Relationship', NS)
        }

        sheets = workbook.findall('main:sheets/main:sheet', NS)
        for sh in sheets:
            title = sh.attrib.get('name', '')
            rid = sh.attrib.get(f"{{{NS['rel']}}}id")
            if not title or not rid:
                continue

            sheet_target = rel_map.get(rid)
            if not sheet_target:
                continue

            sheet_path = _join('xl/workbook.xml', sheet_target)
            sheet_rels_path = posixpath.join(posixpath.dirname(sheet_path), '_rels', posixpath.basename(sheet_path) + '.rels')
            sheet_rels = read_xml(sheet_rels_path)
            if sheet_rels is None:
                continue

            drawing_targets = [
                rel.attrib.get('Target')
                for rel in sheet_rels.findall('pkg:Relationship', NS)
                if 'drawing' in rel.attrib.get('Type', '')
            ]

            items: List[dict] = []
            for drawing_target in drawing_targets:
                drawing_path = _join(sheet_path, drawing_target)
                drawing_xml = read_xml(drawing_path)
                if drawing_xml is None:
                    continue

                drawing_rels_path = posixpath.join(posixpath.dirname(drawing_path), '_rels', posixpath.basename(drawing_path) + '.rels')
                drawing_rels = read_xml(drawing_rels_path)
                if drawing_rels is None:
                    continue

                img_rel_map = {
                    rel.attrib.get('Id'): rel.attrib.get('Target')
                    for rel in drawing_rels.findall('pkg:Relationship', NS)
                }

                anchors = drawing_xml.findall('xdr:oneCellAnchor', NS) + drawing_xml.findall('xdr:twoCellAnchor', NS)
                for idx, anchor in enumerate(anchors, start=1):
                    from_row = anchor.find('xdr:from/xdr:row', NS)
                    row = int(from_row.text) + 1 if from_row is not None and from_row.text is not None else 1

                    blip = anchor.find('.//a:blip', NS)
                    if blip is None:
                        continue
                    embed = blip.attrib.get(f"{{{NS['rel']}}}embed")
                    if not embed:
                        continue

                    media_target = img_rel_map.get(embed)
                    if not media_target:
                        continue

                    media_path = _join(drawing_path, media_target)
                    try:
                        data = zf.read(media_path)
                    except KeyError:
                        continue

                    ext = media_path.rsplit('.', 1)[-1].lower() if '.' in media_path else 'png'
                    if ext not in {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'}:
                        ext = 'png'

                    out_name = f"{chapter_slug}_{safe_slug(title)}_{idx}.{ext}"
                    out_path = ASSET_DIR / out_name
                    out_path.write_bytes(data)

                    items.append({'row': row, 'path': f"/assets/questions/{out_name}"})

            if items:
                result[title] = items

    return result


def assign_images_to_questions(q_map: Dict, sheet_images: List[dict]) -> Dict[int, str]:
    """
    画像を問題に1対1で割り当てる。
    - 画像の行位置の「直上にある問題」に割り当てる
    - 問題と画像の行距離が MAX_ROW_DISTANCE を超える場合は割り当てない
    - 1つの画像は1問にしか割り当てない
    """
    if not sheet_images or not q_map:
        return {}

    MAX_ROW_DISTANCE = 20

    # 問題を行番号順にソート
    sorted_q = sorted((int(v['row']), k) for k, v in q_map.items())

    assignments: Dict[int, str] = {}

    for img in sheet_images:
        img_row = img['row']

        # 画像の行以下にある問題の中で最大行番号のもの（直上の問題）を探す
        best_qno = None
        best_q_row = -1
        for q_row, qno in sorted_q:
            if q_row <= img_row and q_row > best_q_row:
                best_q_row = q_row
                best_qno = qno

        if best_qno is None:
            continue
        if img_row - best_q_row > MAX_ROW_DISTANCE:
            continue
        # 既に割り当て済みの問題には上書きしない（最初の画像を優先）
        if best_qno not in assignments:
            assignments[best_qno] = img['path']

    return assignments


def main():
    all_questions = []
    qid = 1

    for xlsx in sorted(DOC_DIR.glob('*.xlsx')):
        wb = load_workbook(xlsx, data_only=True)
        category = xlsx.stem.replace('第', '第').replace('章', '章 ').strip()
        chapter_slug = safe_slug(xlsx.stem)

        images_by_sheet = extract_images_from_xlsx(xlsx, chapter_slug)

        for ws in wb.worksheets:
            if ws.title in SKIP_SHEETS:
                continue

            q_map, a_map = parse_sheet(ws)
            if not q_map:
                continue

            sheet_images = images_by_sheet.get(ws.title, [])
            image_assignments = assign_images_to_questions(q_map, sheet_images)

            for no in sorted(q_map.keys()):
                meta = q_map[no]
                text = str(meta['text'])
                q_row = int(meta['row'])
                ans = a_map.get(no, '').strip()
                if not ans:
                    continue

                parsed = parse_numeric_answer(ans)
                if parsed:
                    value, unit, abs_tol, rel_tol = parsed
                    question_type = 'numeric'
                    choices = []
                    answer_idx = 0
                    numeric_value = value
                    numeric_unit = unit
                    numeric_abs_tol = abs_tol
                    numeric_rel_tol = rel_tol
                else:
                    question_type = 'choice'
                    built = create_numeric_choices(ans)
                    if built is None:
                        choices, answer_idx = create_fallback_choices(ans)
                    else:
                        choices, answer_idx = built
                    numeric_value = None
                    numeric_unit = ''
                    numeric_abs_tol = None
                    numeric_rel_tol = None

                item = {
                    'id': qid,
                    'category': category,
                    'chapter': ws.title,
                    'difficulty': min(3, 1 + (no - 1) // 3),
                    'text': text,
                    'question_type': question_type,
                    'choices': choices,
                    'answer': answer_idx,
                    'explanation': f'Excel解答: {ans}',
                    'tags': [xlsx.stem, ws.title],
                }

                if numeric_value is not None:
                    item['numeric_answer'] = numeric_value
                    item['numeric_unit'] = numeric_unit
                    item['numeric_tolerance_abs'] = numeric_abs_tol
                    item['numeric_tolerance_rel'] = numeric_rel_tol

                image_path = image_assignments.get(no)
                if image_path:
                    item['image'] = image_path

                all_questions.append(item)
                qid += 1

    js_output = 'const questions = ' + json.dumps(all_questions, ensure_ascii=False, indent=2) + ';\n'
    json_output = json.dumps(all_questions, ensure_ascii=False, indent=2)

    # denki-keisan に出力
    OUT_JS_FILE.write_text(js_output, encoding='utf-8')
    OUT_JSON_FILE.write_text(json_output, encoding='utf-8')
    print(f'Generated {len(all_questions)} questions -> {OUT_JSON_FILE}')

    # denki にも出力（画像パスを /assets/questions/ に変換）
    if DENKI_ROOT.exists():
        js_denki = js_output.replace(
            '"/assets/questions/', '"/assets/questions/'
        )
        DENKI_OUT_JS.parent.mkdir(parents=True, exist_ok=True)
        DENKI_OUT_JS.write_text(js_denki, encoding='utf-8')

        # 画像ファイルも同期
        if DENKI_ASSET_DIR.exists():
            import shutil
            for src_file in ASSET_DIR.glob('*.png'):
                dst = DENKI_ASSET_DIR / src_file.name
                if not dst.exists() or src_file.stat().st_mtime > dst.stat().st_mtime:
                    shutil.copy2(src_file, dst)
        print(f'Also written -> {DENKI_OUT_JS}')


if __name__ == '__main__':
    main()
